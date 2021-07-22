import openpyxl as xl

wb = xl.load_workbook('vgsalesTest.xlsx')
sheet = wb['']
platform = [0]
genre = [0]
publisher = [0]

for row in range(2, sheet.max_row + 1):
    for col in range(3, 6):
        cell = sheet.cell(row, col)
        if col == 3: # Platform
            if platform.count(cell.value) == 0:
                platform.append(cell.value)
            cell.value = platform.index(cell.value)
        elif col == 4: # Genre
            if genre.count(cell.value) == 0:
                genre.append(cell.value)
            cell.value = genre.index(cell.value)
        elif col == 5: # Publisher
            if publisher.count(cell.value) == 0:
                publisher.append(cell.value)
            cell.value = publisher.index(cell.value)

wb.save('vgsalesTest.xlsx')


